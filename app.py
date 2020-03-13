from flask import Flask, render_template, request, send_file, redirect
from forms import Input
from utils import Results, extract_url, getDomainInfo
from actions import submit
from openpyxl import Workbook
from urllib2 import urlopen
from flask_sqlalchemy import SQLAlchemy
import os

SQLALCHEMY_DATABASE_URI = os.environ.get('SQLALCHEMY_DATABASE_URI', None)
SECRET_KEY = os.environ.get('SECRET_KEY', None)
SQLALCHEMY_TRACK_MODIFICATIONS = os.environ.get('SQLALCHEMY_TRACK_MODIFICATIONS', None)


app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = SQLALCHEMY_DATABASE_URI
app.config['SECRET_KEY'] = SECRET_KEY
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = SQLALCHEMY_TRACK_MODIFICATIONS

port = int(os.environ.get("PORT", 5000))

db = SQLAlchemy(app)

class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    url = db.Column(db.String(100), unique=True)
    verdict = db.Column(db.String(20))

    def __init__(self, url, verdict):
        self.url = url
        self.verdict = verdict
       

@app.route('/', methods=['GET', 'POST'])
@app.route('/home', methods=['GET', 'POST'])
def index():
    form = Input(request.form)

    if request.method == 'POST' and form.validate():
        url = form.url.data;        
        base_url = extract_url(url)
        domain_info = getDomainInfo(base_url)['WhoisRecord']
        useful_domain_info = {}
        empty = False

        try:
            del domain_info['registrant']['rawText']
            domain_info['registrant']['street'] = domain_info['registrant'].pop('street1')

            useful_domain_info = {
                'Registrar Name': domain_info['registrarName'],
                'Registrant Details': domain_info['registrant'],
                'Creation Date': domain_info['createdDate'],
                'Updation Date': domain_info['updatedDate'],
                'Expiration Date': domain_info['expiresDate'],
                'Domain Name': base_url
            }


        except KeyError as e:
            if e.message == 'registrant':
                print 'REGISTRANT-------------------------------------------'
                useful_domain_info = {}
                empty = True
            elif e.message == 'street1':
                print 'STREET-------------------------------------------'
                useful_domain_info = {
                    'Registrar Name': domain_info['registrarName'],
                    'Registrant Details': domain_info['registrant'],
                    'Creation Date': domain_info['createdDate'],
                    'Updation Date': domain_info['updatedDate'],
                    'Expiration Date': domain_info['expiresDate'],
                    'Domain Name': base_url,
                }
                empty = False
            else:
                print('---------------ELSE------')
                print e.message    
        except Exception as e:
            print('-----------EXCEPTION-------------')        
            print(e.message)   

        result = submit(url)
        verdict = ''

        if result == Results.SAFE:
            verdict = 'SAFE'
        elif result == Results.MALICIOUS:
            verdict = 'MALICIOUS'
        else:
            verdict = 'MALWARE'

        if db.session.query(Store).filter(Store.url == base_url).count() == 0:
            info = Store(base_url, verdict)
            db.session.add(info)
            db.session.commit()

        if result == Results.SAFE:
            print('----------- SAFE -----------')
            return render_template('safe.html', url=url, base_url=extract_url(url), info=useful_domain_info, isempty=empty)
        elif result == Results.MALICIOUS or result == Results.MALWARE:
            print('----------- MALICIOUS -----------')
            return render_template('malicious.html', url=url, base_url=extract_url(url), info = useful_domain_info, empty=empty)
       
    return render_template('index.html',form=form)


@app.route('/about')
def about():
    return render_template('about.html', title='About')   

@app.route('/download', methods=['POST'])
def download():
    wb = Workbook()

    ws = wb.active
    ws.title = 'Malicious links'

    ws['A1'] = 'Url'
    ws['A1'].style = 'Headline 1'
    ws['B1'] = 'Verdict'
    ws['B1'].style = 'Headline 1'

    results = Store.query.all()
    total_results = len(results) + 1
    index = 0

    for row in ws.iter_rows(min_row = 2, max_col=2, max_row=total_results):
        link = results[index]
        x = 0
        for cell in row:
            if x%2 == 0:
                cell.value = str(link.url)
            else:
                cell.value = str(link.verdict)
            x += 1    
        index += 1    
                
    
    wb.save('static/test.xlsx')  
    print('-------------DATA WRITTEN-----------')
    return send_file('static/test.xlsx')

     

if __name__ == '__main__':
    import main
    app.run(host='0.0.0.0', port=port, debug=True)